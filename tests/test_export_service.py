# 导出服务单元测试

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.services.export_service import ExportService


@pytest.fixture
def db_session():
    """创建内存 SQLite 数据库测试会话"""
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


def test_export_includes_manual_status(db_session):
    """手动匹配 (manual) 记录应计入已匹配"""
    from app.models import Customer, MatchResult, OurReceipt, CustomerSettlement

    # 创建测试数据
    customer = Customer(name="测试客户", slug="test", is_active=True)
    db_session.add(customer)
    db_session.flush()

    receipt = OurReceipt(receipt_no="R001", customer_id=customer.id, period="202608", model="M1", quantity=1, amount=100)
    settlement = CustomerSettlement(customer_id=customer.id, period="202608", match_key="K001", model="M1", quantity=1, amount=100)
    db_session.add(receipt)
    db_session.add(settlement)
    db_session.flush()

    # 手动匹配记录
    manual = MatchResult(
        customer_id=customer.id, period="202608",
        receipt_id=receipt.id, settlement_id=settlement.id,
        match_type="manual", confidence=1.0,
        status="manual", source="manual",
        diff_amount=0, diff_quantity=0,
    )
    db_session.add(manual)
    db_session.commit()

    output = ExportService.export_reconciliation(customer.id, "202608", "测试客户", db_session)
    assert isinstance(output, BytesIO)

    # 验证工作表内容
    wb = load_workbook(output)
    assert "对账汇总" in wb.sheetnames
    ws = wb["对账汇总"]
    rows = list(ws.iter_rows(values_only=True))
    # 第 10 行（索引 9）是已匹配数
    matched_row = [r for r in rows if r[0] == "已匹配数"]
    assert len(matched_row) == 1
    assert matched_row[0][1] == 1, f"手动匹配应计入已匹配: {matched_row}"

    assert "匹配明细" in wb.sheetnames
    ws2 = wb["匹配明细"]
    detail_rows = list(ws2.iter_rows(values_only=True))
    assert len(detail_rows) > 1, "匹配明细应有数据行"
    assert detail_rows[1][0] == "manual", f"匹配类型应为 manual: {detail_rows[1]}"


def test_export_decimal_zero_shown(db_session):
    """Decimal 0 值应显示为 0 而非空白"""
    from app.models import Customer, MatchResult, OurReceipt, CustomerSettlement
    from decimal import Decimal

    customer = Customer(name="零值测试", slug="zero", is_active=True)
    db_session.add(customer)
    db_session.flush()

    receipt = OurReceipt(receipt_no="R002", customer_id=customer.id, period="202608", model="M2", quantity=Decimal("0"), amount=Decimal("0"))
    settlement = CustomerSettlement(customer_id=customer.id, period="202608", match_key="K002", model="M2", quantity=Decimal("0"), amount=Decimal("0"))
    db_session.add(receipt)
    db_session.add(settlement)
    db_session.flush()

    matched = MatchResult(
        customer_id=customer.id, period="202608",
        receipt_id=receipt.id, settlement_id=settlement.id,
        match_type="exact", confidence=Decimal("0.00"), status="matched",
        source="auto", diff_amount=Decimal("0.00"), diff_quantity=Decimal("0.00"),
    )
    db_session.add(matched)
    db_session.commit()

    output = ExportService.export_reconciliation(customer.id, "202608", "零值测试", db_session)
    assert isinstance(output, BytesIO)

    wb = load_workbook(output)
    # 匹配明细 - 零值应显示
    ws = wb["匹配明细"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1]

    # 置信度列（索引 1）
    confidence_val = data[1]
    assert confidence_val is not None and confidence_val != "", f"置信度 0 不应为空白: {confidence_val}"

    # 金额差异明细 - 零差异应包含
    ws5 = wb["金额差异明细"]
    diff_rows = list(ws5.iter_rows(values_only=True))
    assert len(diff_rows) > 1, "金额差异明细应有行（含零差异）"
    assert diff_rows[1][5] == 0, f"金额差异应为 0: {diff_rows[1]}"


def test_export_empty_data_has_consistent_columns(db_session):
    """无数据时工作表应保持列结构一致"""
    from app.models import Customer

    customer = Customer(name="空数据", slug="empty", is_active=True)
    db_session.add(customer)
    db_session.commit()

    output = ExportService.export_reconciliation(customer.id, "202608", "空数据", db_session)
    assert isinstance(output, BytesIO)

    wb = load_workbook(output)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 1, f"{sheet_name} 应至少有表头"
        # 表头不应为空
        header = rows[0]
        assert any(cell is not None for cell in header), f"{sheet_name} 表头不应全空"


def test_export_missing_settlement_no_keyerror(db_session):
    """settlement_id 引用缺失时不抛 KeyError"""
    from app.models import Customer, MatchResult, OurReceipt

    customer = Customer(name="缺失引用", slug="missing", is_active=True)
    db_session.add(customer)
    db_session.flush()

    receipt = OurReceipt(receipt_no="R003", customer_id=customer.id, period="202608", model="M3", quantity=1, amount=100)
    db_session.add(receipt)
    db_session.flush()

    # 匹配结果引用不存在的 settlement_id
    broken = MatchResult(
        customer_id=customer.id, period="202608",
        receipt_id=receipt.id, settlement_id=99999,
        match_type="exact", confidence=1.0, status="matched",
        source="auto", diff_amount=10,
    )
    db_session.add(broken)
    db_session.commit()

    # 不应抛出 KeyError
    output = ExportService.export_reconciliation(customer.id, "202608", "缺失引用", db_session)
    assert isinstance(output, BytesIO)


def test_get_history_includes_manual(db_session):
    """历史查询应计入 manual 状态"""
    from app.models import Customer, MatchResult, OurReceipt, CustomerSettlement

    customer = Customer(name="历史测试", slug="hist", is_active=True)
    db_session.add(customer)
    db_session.flush()

    receipt = OurReceipt(receipt_no="R004", customer_id=customer.id, period="202608", model="M4")
    db_session.add(receipt)
    db_session.flush()

    manual = MatchResult(
        customer_id=customer.id, period="202608",
        receipt_id=receipt.id, match_type="manual",
        confidence=1.0, status="manual", source="manual",
    )
    db_session.add(manual)
    db_session.commit()

    items = ExportService.get_history(db_session)
    assert len(items) == 1
    assert items[0]["matched_count"] == 1, f"manual 应计入 matched_count: {items[0]}"


def test_get_history_empty_customer_ids(db_session):
    """无匹配结果时 get_history 不应报错"""
    items = ExportService.get_history(db_session)
    assert items == []


def _seed_lopsided_results(db_session):
    """构造：2 笔已匹配结算单 + 3 笔仅我方未匹配（结算单侧全部匹配）。

    旧口径（分母=我方笔数）匹配率 2/5=40%；
    正确口径（分母=结算单数）匹配率 2/2=100%。
    """
    from app.models import Customer, MatchResult, OurReceipt, CustomerSettlement

    customer = Customer(name="口径测试", slug="rate", is_active=True)
    db_session.add(customer)
    db_session.flush()

    for i in range(1, 6):
        r = OurReceipt(receipt_no=f"R{i}", customer_id=customer.id, period="202608",
                       model="M", quantity=1, amount=100)
        db_session.add(r)
    db_session.flush()

    receipts = db_session.query(OurReceipt).all()
    # 2 笔匹配（含结算单），3 笔仅我方未匹配
    for i in range(2):
        s = CustomerSettlement(customer_id=customer.id, period="202608",
                               match_key=f"K{i}", model="M", quantity=1, amount=100)
        db_session.add(s)
        db_session.flush()
        db_session.add(MatchResult(customer_id=customer.id, period="202608",
                                   receipt_id=receipts[i].id, settlement_id=s.id,
                                   match_type="exact", confidence=1.0, status="matched",
                                   source="auto", diff_amount=0, diff_quantity=0))
    for i in range(2, 5):
        db_session.add(MatchResult(customer_id=customer.id, period="202608",
                                   receipt_id=receipts[i].id, settlement_id=None,
                                   match_type="未匹配", confidence=0.0, status="unmatched",
                                   source="auto"))
    db_session.commit()
    return customer


def test_export_match_rate_is_settlement_driven(db_session):
    """导出汇总的匹配率以结算单为口径（2/2=100%，而非 2/5=40%）"""
    customer = _seed_lopsided_results(db_session)
    output = ExportService.export_reconciliation(customer.id, "202608", "口径测试", db_session)

    wb = load_workbook(output)
    rows = list(wb["对账汇总"].iter_rows(values_only=True))
    rate_row = [r for r in rows if r[0] == "匹配率"]
    assert len(rate_row) == 1
    assert rate_row[0][1] == "100.0%", f"匹配率应为结算单口径 100.0%: {rate_row}"


def test_get_history_match_rate_is_settlement_driven(db_session):
    """历史查询的匹配率以结算单为口径（2/2=100%，而非 2/5=40%）"""
    customer = _seed_lopsided_results(db_session)
    items = ExportService.get_history(db_session, customer_id=customer.id)
    assert len(items) == 1
    assert items[0]["matched_count"] == 2
    assert items[0]["unmatched_receipts"] == 3
    assert items[0]["match_rate"] == 100.0, f"匹配率应为结算单口径 100.0: {items[0]}"